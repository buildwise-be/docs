"""
Export all help portal content to Excel (reviewer-friendly).

Strips all Mintlify components and markdown syntax so only readable
text remains. Produces one .xlsx file with:
  - One sheet per content section (welcome, support, legal)
    with title, description, seo_title, seo_description, and body text.
  - A "glossary" sheet from config/glossary.yaml
  - A "translations" sheet from config/translations.yaml

Requirements:  pip install openpyxl pyyaml
"""

import re
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DOCS_DIR = Path(__file__).resolve().parent
LANGUAGES = ["en", "nl", "fr", "de"]
SECTIONS = ["welcome", "support", "legal"]
CONTENT_FIELDS = [
    "id",
    "slug",
    "title",
    "description",
    "seo_title",
    "seo_description",
    "section",
    "order",
    "source_language",
    "translation_status",
    "last_updated",
    "source_version",
    "translator",
    "toc",
    "prev_page",
    "next_page",
]

HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PAGE_ID_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
PAGE_ID_FONT = Font(bold=True, size=11)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def parse_frontmatter(text: str) -> tuple[dict, str]:
    """Return (frontmatter_dict, body_markdown) from a .md file's text."""
    match = re.match(r"^---\s*\n(.*?)\n---\s*\n?(.*)", text, re.DOTALL)
    if not match:
        return {}, text
    raw_yaml = re.sub(r"#[^\n]*", "", match.group(1))
    fm = yaml.safe_load(raw_yaml) or {}
    return fm, match.group(2).strip()


def strip_to_plain_text(md: str) -> str:
    """Turn markdown + Mintlify components into plain readable text."""
    text = md

    # Remove self-closing tags like <video ... />
    text = re.sub(r"<video[^>]*/>", "", text)
    # Remove <video ...>...</video> blocks
    text = re.sub(r"<video[^>]*>.*?</video>", "", text, flags=re.DOTALL)
    # Remove full-line HTML-style tags (opening or self-closing)
    text = re.sub(r"<video\b[^>]*>", "", text)

    # Keep text content inside components that wrap readable text.
    # E.g. <Card title="Foo">Some text</Card> -> Foo\nSome text
    def replace_with_title(m):
        title_match = re.search(r'title="([^"]*)"', m.group(1))
        title = title_match.group(1) if title_match else ""
        inner = m.group(2).strip()
        parts = [p for p in (title, inner) if p]
        return "\n".join(parts)

    text = re.sub(
        r"<(Card|Accordion|Step|Tab)\b([^>]*)>(.*?)</\1>",
        lambda m: replace_with_title(re.match(r"(<\w+\b)(.*)$", m.group(0), re.DOTALL) or m),
        text,
        flags=re.DOTALL,
    )
    # More robust: handle Card/Accordion with title attr
    text = re.sub(
        r'<(?:Card|Accordion|Step|Tab)\b[^>]*title="([^"]*)"[^>]*>(.*?)</(?:Card|Accordion|Step|Tab)>',
        lambda m: f"{m.group(1)}\n{m.group(2).strip()}",
        text,
        flags=re.DOTALL,
    )

    # Remove wrapper-only tags (no readable content of their own)
    for tag in [
        "CardGroup", "AccordionGroup", "Steps", "Tabs",
        "Frame", "Note", "Tip", "Warning", "Info", "Check",
    ]:
        text = re.sub(rf"</?{tag}\b[^>]*>", "", text)

    # Remove any remaining HTML/JSX tags
    text = re.sub(r"</?[A-Za-z][A-Za-z0-9]*\b[^>]*>", "", text)

    # Markdown links [text](url) -> text
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    # Markdown images ![alt](url) -> (remove)
    text = re.sub(r"!\[[^\]]*\]\([^)]+\)", "", text)

    # Bold / italic markers
    text = re.sub(r"\*{1,3}([^*]+)\*{1,3}", r"\1", text)
    text = re.sub(r"_{1,3}([^_]+)_{1,3}", r"\1", text)

    # Headings: ## Title -> Title
    text = re.sub(r"^#{1,6}\s+", "", text, flags=re.MULTILINE)

    # Horizontal rules
    text = re.sub(r"^-{3,}\s*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"^\*{3,}\s*$", "", text, flags=re.MULTILINE)

    # Markdown table rows -> keep cell text, join with " | "
    def table_row_to_text(m):
        row = m.group(0)
        if re.match(r"^\|[-\s|:]+\|$", row.strip()):
            return ""
        cells = [c.strip() for c in row.strip().strip("|").split("|")]
        return " | ".join(cells)

    text = re.sub(r"^\|.+\|$", table_row_to_text, text, flags=re.MULTILINE)

    # DEV comments
    text = re.sub(r"^\*To DEVs:.*$", "", text, flags=re.MULTILINE)

    # Collapse multiple blank lines
    text = re.sub(r"\n{3,}", "\n\n", text)

    return text.strip()


def discover_pages(section: str) -> list[str]:
    """Return sorted list of page ids (file stems) for a section."""
    section_dir = DOCS_DIR / "en" / section
    if not section_dir.is_dir():
        return []
    pages = [f.stem for f in sorted(section_dir.glob("*.md"))]
    return pages


def load_page(lang: str, section: str, page_id: str) -> tuple[dict, str]:
    """Load frontmatter + body for a specific lang/section/page."""
    path = DOCS_DIR / lang / section / f"{page_id}.md"
    if not path.exists():
        return {}, ""
    return parse_frontmatter(path.read_text(encoding="utf-8"))


def style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, min_width=12, max_width=60):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        longest = min_width
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                length = max(len(line) for line in lines)
                longest = max(longest, min(length + 2, max_width))
        ws.column_dimensions[col_letter].width = longest


def write_section_sheet(wb: Workbook, section: str):
    ws = wb.create_sheet(title=section.capitalize())
    headers = ["Page ID", "Field"] + [lang.upper() for lang in LANGUAGES]
    ws.append(headers)
    style_header(ws, len(headers))

    pages = discover_pages(section)
    fm_by_lang: dict[str, dict] = {}
    body_by_lang: dict[str, str] = {}

    for page_id in pages:
        fm_by_lang.clear()
        body_by_lang.clear()
        for lang in LANGUAGES:
            fm, body = load_page(lang, section, page_id)
            fm_by_lang[lang] = fm
            body_by_lang[lang] = body

        first_row = True
        for field in CONTENT_FIELDS:
            values = [str(fm_by_lang.get(lang, {}).get(field, "")) for lang in LANGUAGES]
            row = [page_id if first_row else "", field] + values
            ws.append(row)
            r = ws.max_row
            if first_row:
                ws.cell(row=r, column=1).fill = PAGE_ID_FILL
                ws.cell(row=r, column=1).font = PAGE_ID_FONT
                first_row = False
            for c in range(3, 3 + len(LANGUAGES)):
                ws.cell(row=r, column=c).alignment = WRAP_ALIGNMENT

        plain = [strip_to_plain_text(body_by_lang.get(lang, "")) for lang in LANGUAGES]
        body_row = ["", "body"] + plain
        ws.append(body_row)
        r = ws.max_row
        for c in range(3, 3 + len(LANGUAGES)):
            ws.cell(row=r, column=c).alignment = WRAP_ALIGNMENT

        ws.append([])  # blank separator row

    auto_width(ws)
    ws.freeze_panes = "C2"


def write_glossary_sheet(wb: Workbook):
    path = DOCS_DIR / "config" / "glossary.yaml"
    if not path.exists():
        return
    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    terms = data.get("terms", [])

    ws = wb.create_sheet(title="Glossary")
    headers = ["Term ID", "Field"] + [lang.upper() for lang in LANGUAGES]
    ws.append(headers)
    style_header(ws, len(headers))

    for term in terms:
        tid = term.get("id", "")
        term_row = [tid, "term"] + [term.get(lang, "") for lang in LANGUAGES]
        ws.append(term_row)
        r = ws.max_row
        ws.cell(row=r, column=1).fill = PAGE_ID_FILL
        ws.cell(row=r, column=1).font = PAGE_ID_FONT

        desc = term.get("description", {})
        desc_row = ["", "description"] + [desc.get(lang, "") for lang in LANGUAGES]
        ws.append(desc_row)
        for c in range(3, 3 + len(LANGUAGES)):
            ws.cell(row=ws.max_row, column=c).alignment = WRAP_ALIGNMENT

        ws.append([])

    auto_width(ws)
    ws.freeze_panes = "C2"


def flatten_translations(node, prefix=""):
    """Yield (dotted_key, {lang: value}) tuples from nested translations.yaml."""
    if isinstance(node, dict):
        if all(lang in node for lang in LANGUAGES):
            yield prefix, {lang: node.get(lang, "") for lang in LANGUAGES}
        else:
            for key, child in node.items():
                new_prefix = f"{prefix}.{key}" if prefix else key
                yield from flatten_translations(child, new_prefix)


def write_translations_sheet(wb: Workbook):
    path = DOCS_DIR / "config" / "translations.yaml"
    if not path.exists():
        return
    data = yaml.safe_load(path.read_text(encoding="utf-8"))

    ws = wb.create_sheet(title="UI Translations")
    headers = ["Key"] + [lang.upper() for lang in LANGUAGES]
    ws.append(headers)
    style_header(ws, len(headers))

    for dotted_key, lang_vals in flatten_translations(data):
        row = [dotted_key] + [lang_vals.get(lang, "") for lang in LANGUAGES]
        ws.append(row)
        for c in range(2, 2 + len(LANGUAGES)):
            ws.cell(row=ws.max_row, column=c).alignment = WRAP_ALIGNMENT

    auto_width(ws)
    ws.freeze_panes = "B2"


def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for section in SECTIONS:
        write_section_sheet(wb, section)

    write_glossary_sheet(wb)
    write_translations_sheet(wb)

    output = DOCS_DIR / "help_portal_export.xlsx"
    wb.save(output)
    print(f"Exported to {output}")


if __name__ == "__main__":
    main()
