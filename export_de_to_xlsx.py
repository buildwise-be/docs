"""
Export the German help portal Markdown files to XLSX.

The export intentionally keeps the source Markdown as literal as possible:
frontmatter, body Markdown, and the full raw file content are written to separate
columns without stripping Mintlify components or Markdown syntax.

Run:
  python export_de_to_xlsx.py
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DOCS_DIR = Path(__file__).resolve().parent
LOCALE = "de"
OUTPUT_FILE = DOCS_DIR / "german_help_portal_export.xlsx"
SECTIONS = ["welcome", "support", "legal"]
METADATA_FIELDS = [
    "id",
    "slug",
    "locale",
    "title",
    "description",
    "seo_title",
    "seo_description",
    "section",
    "order",
    "source_language",
    "translation_status",
    "last_updated",
    "source_version",
    "translator",
    "toc",
    "prev_page",
    "next_page",
]
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def parse_frontmatter(text: str) -> tuple[str, str, dict[str, str]]:
    """Return raw frontmatter, raw body Markdown, and simple metadata values."""
    match = re.match(r"^---\s*\n(.*?)\n---\s*\n?(.*)", text, re.DOTALL)
    if not match:
        return "", text, {}

    raw_frontmatter = match.group(1)
    body_markdown = match.group(2)
    metadata: dict[str, str] = {}

    for line in raw_frontmatter.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or ":" not in stripped:
            continue

        key, value = stripped.split(":", 1)
        value = value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        metadata[key.strip()] = value

    return raw_frontmatter, body_markdown, metadata


def section_sort_index(path: Path) -> int:
    try:
        return SECTIONS.index(path.parent.name)
    except ValueError:
        return len(SECTIONS)


def order_sort_value(metadata: dict[str, str]) -> int:
    try:
        return int(metadata.get("order", "0"))
    except ValueError:
        return 0


def discover_german_pages() -> list[tuple[Path, str, str, dict[str, str]]]:
    rows = []
    for path in (DOCS_DIR / LOCALE).rglob("*.md"):
        raw_markdown = path.read_text(encoding="utf-8")
        raw_frontmatter, body_markdown, metadata = parse_frontmatter(raw_markdown)
        rows.append((path, raw_frontmatter, body_markdown, metadata))

    return sorted(
        rows,
        key=lambda row: (
            section_sort_index(row[0]),
            order_sort_value(row[3]),
            row[0].as_posix(),
        ),
    )


def auto_width(ws, min_width: int = 12, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        longest = min_width

        for cell in col_cells:
            if cell.value is None:
                continue

            lines = str(cell.value).split("\n")
            length = max(len(line) for line in lines)
            longest = max(longest, min(length + 2, max_width))

        ws.column_dimensions[col_letter].width = longest


def style_header(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def export_to_xlsx() -> Path:
    pages = discover_german_pages()
    headers = [
        "source_path",
        "file_name",
        *METADATA_FIELDS,
        "raw_frontmatter",
        "body_markdown",
        "raw_markdown",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "German Content"
    ws.append(headers)
    style_header(ws, len(headers))

    for path, raw_frontmatter, body_markdown, metadata in pages:
        raw_markdown = path.read_text(encoding="utf-8")
        row = {
            "source_path": path.relative_to(DOCS_DIR).as_posix(),
            "file_name": path.name,
            "raw_frontmatter": raw_frontmatter,
            "body_markdown": body_markdown,
            "raw_markdown": raw_markdown,
        }
        row.update({field: metadata.get(field, "") for field in METADATA_FIELDS})
        ws.append([row.get(header, "") for header in headers])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT

    auto_width(ws)
    ws.freeze_panes = "C2"
    wb.save(OUTPUT_FILE)

    return OUTPUT_FILE


def main() -> None:
    output = export_to_xlsx()
    print(f"Exported German help portal content to {output}")


if __name__ == "__main__":
    main()
